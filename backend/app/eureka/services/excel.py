from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List
from backend.app.eureka.models import EurekaRegistration

def generate_eureka_excel(registrations: List[EurekaRegistration]) -> BytesIO:
    """
    Generates a stylized Excel sheet containing registrations and team members.
    Returns a BytesIO stream containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Eureka Registrations"
    
    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Styling definitions
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=10)
    lead_font = Font(name=font_family, size=10, bold=True)
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Dark Slate
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")  # Muted Dark Gray
    lead_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")  # Very Light Green for Team Lead
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 1. Title Row
    ws.merge_cells("A1:AP1")
    title_cell = ws["A1"]
    title_cell.value = "EUREKA! Startup Pitching Registrations"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center
    ws.row_dimensions[1].height = 40
    
    # 2. Header Columns
    headers = [
        # Startup basic details
        "Reg ID", "Startup Name", "Category", "Current Stage", "Description", 
        "Problem Statement", "Solution", "Existing Startup?", "Website", 
        "Current Stage Details", "Team Size", "Revenue", "Registration Details", 
        "Pitch Deck Uploaded?", "Status", "Submission Date"
    ]
    
    # We support up to 5 team members in columns
    for i in range(1, 6):
        prefix = f"Team Lead (M1)" if i == 1 else f"Member {i}"
        headers.extend([
            f"{prefix} Name", f"{prefix} Email", f"{prefix} Phone", 
            f"{prefix} College", f"{prefix} Department", f"{prefix} Year", f"{prefix} Bank Account"
        ])
        
    ws.row_dimensions[2].height = 30
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # 3. Data Insertion
    current_row = 3
    for reg in registrations:
        ws.row_dimensions[current_row].height = 24
        
        # Format registration values
        reg_vals = [
            reg.registration_id,
            reg.startup_name,
            reg.category,
            reg.stage,
            reg.description,
            reg.problem_statement,
            reg.solution,
            "Yes" if reg.is_existing else "No",
            reg.website or "N/A",
            reg.current_stage or "N/A",
            reg.team_size,
            reg.revenue or "N/A",
            reg.registration_details or "N/A",
            "Yes" if reg.has_pitch_deck else "No",
            reg.status,
            reg.created_at.strftime("%Y-%m-%d %H:%M")
        ]
        
        # Sort team members so that Lead is first
        sorted_members = sorted(reg.team_members, key=lambda x: not x.is_lead)
        
        # Populate startup columns
        for col_idx, val in enumerate(reg_vals, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = align_center if col_idx in [1, 3, 4, 8, 11, 14, 15, 16] else align_left
            
            # Status styling colors
            if col_idx == 15: # Status column
                if val == "Approved":
                    cell.fill = PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid") # light green
                    cell.font = Font(name=font_family, size=10, bold=True, color="03543F")
                elif val == "Rejected":
                    cell.fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid") # light red
                    cell.font = Font(name=font_family, size=10, bold=True, color="9B1C1C")
                else:
                    cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # light yellow
                    cell.font = Font(name=font_family, size=10, bold=True, color="713F12")
        
        # Populate team member columns
        member_col_offset = 17
        for member_idx in range(5):
            if member_idx < len(sorted_members):
                m = sorted_members[member_idx]
                m_vals = [m.name, m.email, m.phone, m.college, m.department or "", m.year or "", m.bank_account or ""]
            else:
                m_vals = ["", "", "", "", "", "", ""]
                
            for detail_idx, val in enumerate(m_vals):
                col_number = member_col_offset + (member_idx * 7) + detail_idx
                cell = ws.cell(row=current_row, column=col_number, value=val)
                cell.font = lead_font if member_idx == 0 else body_font
                cell.border = thin_border
                cell.alignment = align_left
                
                # Light highlight for Team Lead column block
                if member_idx == 0 and val != "":
                    cell.fill = lead_fill
                    
        current_row += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Skip the title row (row 1) since it's merged and will stretch columns
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
                
        # Limit column width to avoid ridiculously wide columns for text descriptions
        adjusted_width = min(max(max_len + 3, 10), 30)
        ws.column_dimensions[col_letter].width = adjusted_width
        
    # Write file stream
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream
